"""Human-readable exports (brief 3.2): clients + invoices as .xlsx.

These are read-only copies of the records, written into the `exports\\` folder so
the data can always be opened outside the app. Identifiers (customer codes, GST
and account numbers) are written as text so Excel never reformats them.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from . import config, db
from .utils import nz_date

CLIENT_HEADERS = ["client_id", "customer_code", "name", "address", "phone",
                  "email", "created_date", "notes"]
INVOICE_HEADERS = ["invoice_number", "client_name", "reference", "invoice_date",
                   "details", "subtotal", "gst", "amount_payable", "status",
                   "paid_date", "is_historical"]
# Columns that must stay text so leading zeros / dashes survive in Excel.
_TEXT_COLS = {"customer_code", "phone", "reference"}


def _autosize(ws) -> None:
    for col in ws.columns:
        width = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 60)


def export_clients() -> Path:
    conn = db.get_db()
    rows = conn.execute(
        "SELECT client_id, customer_code, name, address, phone, email, "
        "created_date, notes FROM clients ORDER BY name COLLATE NOCASE"
    ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Clients"
    ws.append(CLIENT_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows:
        ws.append([r[h] for h in CLIENT_HEADERS])
    _force_text(ws, CLIENT_HEADERS)
    _autosize(ws)

    out = config.exports_dir() / "clients.xlsx"
    wb.save(out)
    return out


def export_invoices() -> Path:
    conn = db.get_db()
    rows = conn.execute(
        "SELECT invoice_number, client_name_snapshot, reference, invoice_date, "
        "details, subtotal, gst, amount_payable, status, paid_date, is_historical "
        "FROM invoices ORDER BY invoice_number"
    ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"
    ws.append(INVOICE_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows:
        ws.append([
            r["invoice_number"], r["client_name_snapshot"], r["reference"],
            nz_date(r["invoice_date"]), r["details"], r["subtotal"], r["gst"],
            r["amount_payable"], r["status"], nz_date(r["paid_date"]), r["is_historical"],
        ])
    _force_text(ws, INVOICE_HEADERS)
    _autosize(ws)

    out = config.exports_dir() / "invoices.xlsx"
    wb.save(out)
    return out


def _force_text(ws, headers) -> None:
    """Mark identifier-ish columns as text so Excel leaves them untouched."""
    for idx, name in enumerate(headers, start=1):
        if name in _TEXT_COLS:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=idx).number_format = "@"


def export_all() -> dict[str, Path]:
    return {"clients": export_clients(), "invoices": export_invoices()}
